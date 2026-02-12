from __future__ import annotations

import csv
import io
import json
from collections.abc import Iterable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.entities import Lead
from app.models.schemas import LeadImportResult

_FIELD_ALIASES = {
    "name": {"name", "full_name", "lead_name", "contact_name"},
    "email": {"email", "email_address", "mail"},
    "company": {"company", "company_name", "organization", "org"},
    "position": {"position", "title", "job_title", "role"},
    "niche": {"niche", "segment", "industry"},
}


class LeadImporter:
    def __init__(self, db: Session) -> None:
        self.db = db

    def parse(self, filename: str, payload: bytes) -> list[dict[str, str]]:
        lower = filename.lower()
        if lower.endswith(".csv"):
            return self._parse_csv(payload)
        if lower.endswith(".json"):
            return self._parse_json(payload)
        if lower.endswith(".txt"):
            return self._parse_txt(payload)
        if lower.endswith(".xlsx"):
            return self._parse_xlsx(payload)
        raise ValueError("Unsupported file type. Use CSV, TXT, JSON, or XLSX.")

    def import_rows(self, rows: Iterable[dict[str, str]]) -> LeadImportResult:
        inserted = skipped_existing = skipped_opted_out = invalid_rows = 0
        for raw in rows:
            row = self._normalize_row(raw)
            email = (row.get("email") or "").strip().lower()
            name = (row.get("name") or "").strip()

            if not email or not name:
                invalid_rows += 1
                continue

            existing = self.db.scalar(select(Lead).where(Lead.email == email))
            if existing:
                if existing.opt_out:
                    skipped_opted_out += 1
                else:
                    skipped_existing += 1
                continue

            lead = Lead(
                name=name,
                email=email,
                company=(row.get("company") or "").strip() or None,
                position=(row.get("position") or "").strip() or None,
                niche=(row.get("niche") or "").strip() or None,
            )
            self.db.add(lead)
            inserted += 1

        self.db.commit()
        return LeadImportResult(
            inserted=inserted,
            skipped_existing=skipped_existing,
            skipped_opted_out=skipped_opted_out,
            invalid_rows=invalid_rows,
        )

    def _parse_csv(self, payload: bytes) -> list[dict[str, str]]:
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    def _parse_json(self, payload: bytes) -> list[dict[str, str]]:
        text = payload.decode("utf-8")
        data = json.loads(text)
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list):
            raise ValueError("JSON import expects an object or an array of objects.")
        return [dict(row) for row in data if isinstance(row, dict)]

    def _parse_txt(self, payload: bytes) -> list[dict[str, str]]:
        text = payload.decode("utf-8")
        rows: list[dict[str, str]] = []
        for line in text.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            parts = [p.strip() for p in stripped.split(",")]
            if len(parts) >= 2:
                rows.append(
                    {
                        "name": parts[0],
                        "email": parts[1],
                        "company": parts[2] if len(parts) > 2 else "",
                        "position": parts[3] if len(parts) > 3 else "",
                    }
                )
            else:
                rows.append({"name": "", "email": ""})
        return rows

    def _parse_xlsx(self, payload: bytes) -> list[dict[str, str]]:
        workbook = load_workbook(filename=io.BytesIO(payload), read_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        records: list[dict[str, str]] = []
        for values in rows_iter:
            row: dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = values[idx] if idx < len(values) else None
                row[header] = "" if value is None else str(value).strip()
            if any(v for v in row.values()):
                records.append(row)
        return records

    def _normalize_row(self, row: dict[str, str]) -> dict[str, str]:
        lowered = {(k or "").strip().lower(): (v if isinstance(v, str) else str(v or "")) for k, v in row.items()}
        normalized: dict[str, str] = {}
        for target, aliases in _FIELD_ALIASES.items():
            value = ""
            for alias in aliases:
                if alias in lowered and lowered[alias].strip():
                    value = lowered[alias].strip()
                    break
            normalized[target] = value
        return normalized
